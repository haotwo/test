# -*- coding:utf-8 -*-
#作者 ：Lyle.Li
#时间 :2020/4/16 10:37
#文件 :do_excel.py

from openpyxl import load_workbook
from tools.project_path import *
from tools.read_congig import ReadConfig
from tools.get_data import GetData
class DoExcel:
    @classmethod
    def get_data(cls,file_name):
        wb=load_workbook(file_name)
        mode=eval(ReadConfig.get_config(config_path,'MODE','mode'))
        tel=getattr(GetData,'NoRegName')
        test_data=[]
        for key in mode:   #遍历这个配置的字典
            sheet=wb[key]  #表单名
            if mode[key]=='all':               #对测试配置文件进行判断
                for i in range(2,sheet.max_row+1):
                    row_data={}
                    row_data['case_id']=sheet.cell(i,1).value
                    row_data['url']=sheet.cell(i,2).value
                    # row_data['data']=sheet.cell(i,3).value
                    if sheet.cell(i,3).value.find('${tel}')!=-1:
                        row_data['data']=sheet.cell(i,3).value.replace('${tel}',str(tel))
                    elif sheet.cell(i,3).value.find('${tel_1}')!=-1:
                        row_data['data']=sheet.cell(i,3).value.replace('${tel_1}',str(tel+1))
                    else:
                        row_data['data']=sheet.cell(i,3).value

                    row_data['title']=sheet.cell(i,4).value
                    row_data['http_method']=sheet.cell(i,5).value
                    row_data['expect']=sheet.cell(i,6).value
                    row_data['sheet_name']=key
                    test_data.append(row_data)
                    cls.update_tel(file_name,'init',tel+2)
            else:
                for case_id in mode[key]:
                    row_data={}
                    row_data['case_id']=sheet.cell(case_id+1,1).value
                    row_data['url']=sheet.cell(case_id+1,2).value
                    # row_data['data']=sheet.cell(case_id+1,3).value
                    if sheet.cell(case_id,3).value.find('${tel}')!=-1:
                        row_data['data']=sheet.cell(case_id,3).value.replace('${tel}',str(tel))
                    elif sheet.cell(case_id,3).value.find('${tel_1}')!=-1:
                        row_data['data']=sheet.cell(case_id,3).value.replace('${tel_1}',str(tel+1))
                    else:
                        row_data['data']=sheet.cell(i,3).value
                    row_data['title']=sheet.cell(case_id+1,4).value
                    row_data['http_method']=sheet.cell(case_id+1,5).value
                    row_data['expect']=sheet.cell(case_id+1,6).value
                    row_data['sheet_name']=key
                    test_data.append(row_data)
                    cls.update_tel(file_name,'init',tel+2)


        return test_data

    def write_back(self,file_name,sheet_name,i,result,TestResult):     #写回测试数据
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(i,7).value=result
        sheet.cell(i,8).value=TestResult
        wb.save(file_name)

    @staticmethod
    def update_tel(filename,sheet_name,tel,):       #将测试数据更新excel中
        wb=load_workbook(filename)
        sheet=wb[sheet_name]
        sheet.cell(2,1).value=tel
        wb.save(filename)
if __name__ == '__main__':
    data=DoExcel().get_data(case_path)
    print(data)

